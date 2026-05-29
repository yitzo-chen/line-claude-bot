"""B1 鋼筋混凝土規範查詢模組"""
import os
import openpyxl

_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(_DIR, "data", "regulations", "鋼筋混凝土標準規範.xlsx")
VERSION = "R1A（XF11B-0000-0001）"
VERSION_NOTE = f"\n\n📋 資料版次：{VERSION}\n使用前請確認與施工圖說版次一致，勿以舊版資料施作。"

# 觸發規範查詢的關鍵字
_KEYWORDS = [
    "伸展長度", "發展長度", "搭接長度", "搭接", "彎鉤", "保護層", "淨間距",
    "D10", "D13", "D16", "D19", "D22", "D25", "D29", "D32", "D36",
    "fc'", "fc′", "鋼筋規範", "竹節鋼筋",
    "拉力", "壓力", "頂層筋", "箍筋", "A級", "B級搭",
]


def is_reg_query(text: str) -> bool:
    return any(kw in text for kw in _KEYWORDS)


_ctx_cache: str | None = None


def _load_context() -> str:
    global _ctx_cache
    if _ctx_cache is not None:
        return _ctx_cache

    wb = openpyxl.load_workbook(XLSX)
    parts = []

    # Sheet 1：伸展長度及搭接長度（結構化讀取，跳過合併空值列）
    ws1 = wb.worksheets[0]
    lines = [f"【{ws1.title}】（單位：mm）"]
    # 固定欄位名稱（對應 gen_rebar_xlsx.py 的欄順序）
    col_names = [
        "fc'(kgf/cm²)", "號數",
        "伸展長度-其他(Ld)", "伸展長度-頂層(×1.3)",
        "伸展長度-壓力", "彎鉤伸展-臨界(Ldh)", "彎鉤伸展-其他(Ldh)",
        "搭接-A級(×1.0)", "搭接-B級(×1.3)", "搭接-B級頂層",
        "壓力搭接(Lg)", "db(mm)",
    ]
    lines.append(" | ".join(col_names))
    for row in ws1.iter_rows(min_row=5, values_only=True):
        if row[1] is None:
            continue  # 跳過空列
        vals = [str(v) if v is not None else "-" for v in row]
        lines.append(" | ".join(vals))
    lines.append("備註：頂層鋼筋=澆置深度>300mm以下水平筋；B級搭接=搭接區>50%同時搭接或As提供/As需求<2")
    parts.append("\n".join(lines))

    # Sheet 2：標準彎鉤
    ws2 = wb.worksheets[1]
    lines2 = [f"【{ws2.title}】（單位：mm）"]
    col2 = ["號數", "db(mm)", "180°彎鉤-直徑D", "180°彎鉤-延伸A", "90°彎鉤-直徑D", "90°彎鉤-延伸A", "箍筋-直徑D", "箍筋-延伸A"]
    lines2.append(" | ".join(col2))
    for row in ws2.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        vals = [str(v) if v is not None else "-" for v in row]
        lines2.append(" | ".join(vals))
    parts.append("\n".join(lines2))

    # Sheet 3：保護層
    ws3 = wb.worksheets[2]
    lines3 = [f"【{ws3.title}】（單位：mm）"]
    lines3.append("狀況 | 條件 | 版/牆(mm) | 梁/柱/基腳(mm)")
    for row in ws3.iter_rows(min_row=4, values_only=True):
        vals = [str(v).strip() if v is not None else "" for v in row]
        line = " | ".join(v for v in vals if v)
        if line:
            lines3.append(line)
    parts.append("\n".join(lines3))

    # Sheet 4：最小淨間距
    ws4 = wb.worksheets[3]
    lines4 = [f"【{ws4.title}】"]
    for row in ws4.iter_rows(min_row=2, values_only=True):
        vals = [str(v).strip() if v is not None else "" for v in row]
        line = " | ".join(v for v in vals if v)
        if line:
            lines4.append(line)
    parts.append("\n".join(lines4))

    _ctx_cache = "\n\n".join(parts)
    return _ctx_cache


# 給 ask_groq 用的 system prompt
REG_SYSTEM = (
    "你是台灣工地專用的鋼筋混凝土規範查詢助理，只依據提供的規範資料回答。"
    "回答規則：\n"
    "1. 數值直接列出並標示單位(mm)\n"
    "2. 同時列出「其他鋼筋」和「頂層鋼筋」兩種情況（若相關）\n"
    "3. 若查無對應資料，明確說明\n"
    "4. 不要自行推算或補充規範外資訊\n"
    "5. 繁體中文，格式簡潔"
)


def build_query_message(user_text: str) -> str:
    """組合含規範資料的完整查詢訊息，傳給 Groq"""
    ctx = _load_context()
    return (
        f"以下是鋼筋混凝土標準規範資料（版次 {VERSION}）：\n\n"
        f"{ctx}\n\n"
        f"---\n"
        f"使用者問題：{user_text}"
    )
