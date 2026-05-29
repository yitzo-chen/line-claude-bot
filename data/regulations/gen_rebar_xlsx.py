"""
產生鋼筋混凝土標準規範 xlsx
來源：XF11B-0000-0001_R1A_sign.pdf
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── 樣式 ──────────────────────────────────────────────────────────────────────
def border(thick=False):
    s = "medium" if thick else "thin"
    side = Side(style=s)
    return Border(left=side, right=side, top=side, bottom=side)

def hdr(bg="1F4E79", fg="FFFFFF", bold=True, size=9, wrap=True):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(color=fg, bold=bold, size=size, name="微軟正黑體")
    ali  = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return fill, font, ali

def cell_style(ws, row, col, value, fill=None, font=None, ali=None, brdr=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill: c.fill = fill
    if font: c.font = font
    if ali:  c.alignment = ali
    if brdr: c.border = brdr
    return c

NORMAL_FONT = Font(size=9, name="微軟正黑體")
NORMAL_ALI  = Alignment(horizontal="center", vertical="center")
BOLD_FONT   = Font(size=9, bold=True, name="微軟正黑體")
THIN_BORDER = border(False)
THICK_BORDER= border(True)

H1_FILL, H1_FONT, H1_ALI = hdr("1F4E79", "FFFFFF")
H2_FILL, H2_FONT, H2_ALI = hdr("2E75B6", "FFFFFF")
H3_FILL, H3_FONT, H3_ALI = hdr("BDD7EE", "000000")

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1：發展長度及搭接長度
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "伸展長度及搭接長度"

title = "竹節鋼筋伸展及搭接長度（非預力現場澆置混凝土）  單位：mm"
ws1.merge_cells("A1:M1")
c = ws1["A1"]
c.value = title
c.font = Font(size=11, bold=True, name="微軟正黑體")
c.alignment = Alignment(horizontal="center", vertical="center")
c.fill = PatternFill("solid", fgColor="1F4E79")
c.font = Font(size=11, bold=True, name="微軟正黑體", color="FFFFFF")
ws1.row_dimensions[1].height = 22

# ── 欄位標題（第2~4列）─────────────────────────────────────────────────────
headers_main = [
    ("A", 2, 1, "混凝土\nfc'\n(kgf/cm²)"),
    ("B", 2, 2, "鋼筋\n號數"),
]

# 第2列合併標題
merge_def = [
    # (start_row, end_row, start_col, end_col, text, fill)
    (2, 4, 1, 1, "混凝土\nfc'\n(kgf/cm²)", H2_FILL),
    (2, 4, 2, 2, "鋼筋\n號數", H2_FILL),
    (2, 2, 3, 5, "最小拉壓力伸展長度（Ld）", H1_FILL),
    (2, 2, 6, 7, "標準彎鉤最小\n拉力發展長度（Ldh）", H1_FILL),
    (2, 2, 8,10, "拉力搭接長度", H1_FILL),
    (2, 4,11,11, "壓力\n搭接長度", H2_FILL),
    (2, 4,12,12, "db\n(mm)", H2_FILL),
    (3, 3, 3, 4, "拉力伸展長度", H2_FILL),
    (3, 4, 5, 5, "壓力\n伸展長度\n(Ld)", H3_FILL),
    (3, 4, 6, 6, "臨界\n斷面\n(Ldh)", H3_FILL),
    (3, 4, 7, 7, "其他\n(Ldh)", H3_FILL),
    (3, 3, 8, 9, "搭接長度", H2_FILL),
    (3, 4,10,10, "B級搭接\n(×1.3)", H3_FILL),
    (4, 4, 3, 3, "其他鋼筋\n(Ld)", H3_FILL),
    (4, 4, 4, 4, "頂層鋼筋\n(×1.3)", H3_FILL),
    (4, 4, 8, 8, "A級搭接\n(×1.0)", H3_FILL),
    (4, 4, 9, 9, "B級搭接\n(×1.3)", H3_FILL),
]

for sr, er, sc, ec, text, fill in merge_def:
    ws1.merge_cells(start_row=sr, end_row=er, start_column=sc, end_column=ec)
    c = ws1.cell(row=sr, column=sc, value=text)
    c.fill = fill
    c.font = Font(size=8, bold=True, name="微軟正黑體",
                  color="FFFFFF" if fill in (H1_FILL, H2_FILL) else "000000")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for r in [2,3,4]:
    ws1.row_dimensions[r].height = 28

# ── 資料 ──────────────────────────────────────────────────────────────────────
# 欄：fc' | 號數 | Ld其他 | Ld頂層 | Ld壓力 | Ldh臨界 | Ldh其他 | A級拉 | B級拉 | B級拉頂 | 壓搭 | db

data_210 = [
    #  fc'  号数  Ld其  Ld頂   Ldc   Ldh臨  Ldh其  A拉   B拉   B拉頂  壓搭   db
    (210, "D10",  310,  410,  200,  150,  150,  310,  410,  570,  300,  9.53),
    (210, "D13",  420,  540,  200,  150,  150,  420,  540,  700,  300, 12.70),
    (210, "D16",  520,  680,  220,  150,  150,  520,  680,  880,  300, 15.88),
    (210, "D19",  940, 1220,  420,  150,  150,  940, 1220, 1580,  580, 19.05),
    (210, "D22", 1360, 1770,  490,  490,  490, 1360, 1770, 2290,  670, 22.23),
    (210, "D25", 1660, 2160,  560,  560,  560, 1660, 2160, 2810,  760, 25.40),
    (210, "D29", 1970, 2560,  700,  700,  700, 1970, 2560, 3330,  970, 28.58),
    (210, "D32", 2300, 2840,  780,  780,  780, 2300, 2840, 3700, 1070, 31.75),
    (210, "D36", 2590, 3360,  900,  900,  900, 2590, 3360, 4370, 1270, 34.93),
]

data_280 = [
    (280, "D10",  270,  350,  200,  150,  150,  270,  350,  460,  300,  9.53),
    (280, "D13",  360,  470,  200,  150,  150,  360,  470,  610,  300, 12.70),
    (280, "D16",  450,  590,  200,  150,  150,  450,  590,  760,  320, 15.88),
    (280, "D19",  810, 1060,  200,  150,  150,  810, 1060, 1380,  300, 19.05),
    (280, "D22", 1180, 1530,  420,  420,  420, 1180, 1530, 1990,  580, 22.23),
    (280, "D25", 1350, 1750,  480,  480,  480, 1350, 1750, 2270,  660, 25.40),
    (280, "D29", 1715, 2220,  610,  610,  610, 1715, 2220, 2880,  840, 28.58),
    (280, "D32", 1900, 2460,  680,  680,  680, 1900, 2460, 3200,  930, 31.75),
    (280, "D36", 2140, 2780,  760,  760,  760, 2140, 2780, 3610, 1040, 34.93),
]

ROW_START = 5
FILL_210  = PatternFill("solid", fgColor="DEEAF1")
FILL_280  = PatternFill("solid", fgColor="E2EFDA")

row = ROW_START
for block, fill in [(data_210, FILL_210), (data_280, FILL_280)]:
    first = row
    for rec in block:
        for col, val in enumerate(rec, start=1):
            c = ws1.cell(row=row, column=col, value=val)
            c.font  = NORMAL_FONT
            c.alignment = NORMAL_ALI
            c.fill  = fill
            c.border = THIN_BORDER
        ws1.row_dimensions[row].height = 16
        row += 1
    # 合併 fc' 欄
    ws1.merge_cells(start_row=first, end_row=row-1, start_column=1, end_column=1)
    ws1["A" + str(first)].alignment = Alignment(horizontal="center", vertical="center")

# ── 備註 ──────────────────────────────────────────────────────────────────────
notes = [
    "備註：",
    "1. 頂層鋼筋 = 澆置深度超過 300mm 以下之水平鋼筋，伸展長度 × 1.3",
    "2. B 級搭接 = 搭接區超過 50% 鋼筋同時搭接，或 As提供/As需求 < 2，搭接長度 × 1.3",
    "3. 輕量混凝土伸展長度 × 1.3",
    "4. 環氧樹脂塗層鋼筋伸展長度 × 1.5（與頂層係數疊加最大 × 1.7）",
    "5. 版次：R1A（XF11B-0000-0001_R1A），來源：混凝土結構設計規範",
]
note_row = row + 1
for i, n in enumerate(notes):
    ws1.merge_cells(start_row=note_row+i, end_row=note_row+i, start_column=1, end_column=12)
    c = ws1.cell(row=note_row+i, column=1, value=n)
    c.font = Font(size=8, name="微軟正黑體", bold=(i==0))
    c.alignment = Alignment(horizontal="left", vertical="center")

# ── 欄寬 ──────────────────────────────────────────────────────────────────────
col_widths = [8, 7, 9, 9, 9, 9, 9, 9, 9, 9, 9, 7]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2：標準彎鉤及彎曲規定
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("標準彎鉤及彎曲規定")

ws2.merge_cells("A1:H1")
ws2["A1"].value = "標準彎鉤及彎曲規定（Standard Hooks and Minimum Bend）  單位：mm"
ws2["A1"].font  = Font(size=11, bold=True, name="微軟正黑體", color="FFFFFF")
ws2["A1"].fill  = PatternFill("solid", fgColor="1F4E79")
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 22

hook_headers = [
    (2, 3, 1, 1, "鋼筋號數", H2_FILL),
    (2, 3, 2, 2, "db (mm)", H2_FILL),
    (2, 2, 3, 4, "180° 彎鉤", H1_FILL),
    (2, 2, 5, 6, "90° 彎鉤", H1_FILL),
    (2, 2, 7, 8, "箍筋彎鉤\n(90° 或 135°)", H1_FILL),
    (3, 3, 3, 3, "最小彎曲\n直徑 D", H3_FILL),
    (3, 3, 4, 4, "末端延伸\nA (≥4db, ≥6cm)", H3_FILL),
    (3, 3, 5, 5, "最小彎曲\n直徑 D", H3_FILL),
    (3, 3, 6, 6, "末端延伸\nA (≥12db)", H3_FILL),
    (3, 3, 7, 7, "最小彎曲\n直徑 D", H3_FILL),
    (3, 3, 8, 8, "末端延伸\nA (≥6db/12db)", H3_FILL),
]
for sr, er, sc, ec, text, fill in hook_headers:
    ws2.merge_cells(start_row=sr, end_row=er, start_column=sc, end_column=ec)
    c = ws2.cell(row=sr, column=sc, value=text)
    c.fill = fill
    c.font = Font(size=8, bold=True, name="微軟正黑體",
                  color="FFFFFF" if fill in (H1_FILL, H2_FILL) else "000000")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for r in [2, 3]:
    ws2.row_dimensions[r].height = 30

# 號數 | db | 180°D | 180°A | 90°D | 90°A | 箍D | 箍A
hook_data = [
    ("D10",  9.53,  60,  40,  60,  120,  40, 110),
    ("D13", 12.70,  80,  60,  80,  160,  40, 110),
    ("D16", 15.88, 100,  70, 100,  200,  80, 120),
    ("D19", 19.05, 120,  80, 120,  230,  80, 120),
    ("D22", 22.23, 145, 100, 145,  270, 145, 150),
    ("D25", 25.40, 180, 130, 180,  310, 180, 180),
    ("D29", 28.58, 200, 160, 200,  350, 200, 200),
    ("D32", 31.75, 250, 180, 250,  390, 250, 250),
    ("D36", 34.93, 290, 200, 290,  430, 290, 290),
]
for i, row_data in enumerate(hook_data):
    r = 4 + i
    for col, val in enumerate(row_data, 1):
        c = ws2.cell(row=r, column=col, value=val)
        c.font = NORMAL_FONT
        c.alignment = NORMAL_ALI
        c.border = THIN_BORDER
        c.fill = PatternFill("solid", fgColor="DEEAF1" if i % 2 == 0 else "FFFFFF")
    ws2.row_dimensions[r].height = 16

for i, w in enumerate([8, 8, 10, 14, 10, 14, 10, 16], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3：保護層
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("混凝土保護層")

ws3.merge_cells("A1:E1")
ws3["A1"].value = "混凝土最小保護層厚度（非預力現場澆置）  單位：mm"
ws3["A1"].font  = Font(size=11, bold=True, name="微軟正黑體", color="FFFFFF")
ws3["A1"].fill  = PatternFill("solid", fgColor="1F4E79")
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 22

cover_headers = [
    (2, 3, 1, 1, "狀況", H2_FILL),
    (2, 3, 2, 3, "條件說明", H2_FILL),
    (2, 2, 4, 5, "最小保護層厚度", H1_FILL),
    (3, 3, 4, 4, "版／牆\n(mm)", H3_FILL),
    (3, 3, 5, 5, "梁／柱／基腳\n(mm)", H3_FILL),
]
for sr, er, sc, ec, text, fill in cover_headers:
    ws3.merge_cells(start_row=sr, end_row=er, start_column=sc, end_column=ec)
    c = ws3.cell(row=sr, column=sc, value=text)
    c.fill = fill
    c.font = Font(size=8, bold=True, name="微軟正黑體",
                  color="FFFFFF" if fill in (H1_FILL, H2_FILL) else "000000")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for r in [2, 3]:
    ws3.row_dimensions[r].height = 28

cover_data = [
    # (狀況, 狀況合併, 條件, 版牆, 梁柱基腳)
    ("(1)", True,  "不受風雨侵襲且不與土壤接觸　≤ D16", 20, 40),
    ("",   False, "不受風雨侵襲且不與土壤接觸　≥ D19", 20, 40),
    ("(2)", True,  "受風雨侵襲　≤ D16", 40, 40),
    ("",   False, "受風雨侵襲　≥ D19", 50, 50),
    ("(3)", False, "經常與水或土壤接觸", 65, 75),
    ("(4)", False, "直接澆置於土壤或岩石上", 75, 75),
    ("(5)", False, "與海水或腐蝕性環境接觸", 100, 100),
]

fills = [PatternFill("solid", fgColor="DEEAF1"), PatternFill("solid", fgColor="FFFFFF")]
r = 4
for idx, (num, merge, desc, slab, beam) in enumerate(cover_data):
    for col, val in enumerate([num, desc, desc, slab, beam], 1):
        if col == 2: continue
        c = ws3.cell(row=r, column=col, value=val)
        c.font = NORMAL_FONT
        c.alignment = NORMAL_ALI
        c.border = THIN_BORDER
        c.fill = fills[idx % 2]
    ws3.merge_cells(start_row=r, end_row=r, start_column=2, end_column=3)
    c = ws3.cell(row=r, column=2, value=desc)
    c.font = NORMAL_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = THIN_BORDER
    c.fill = fills[idx % 2]
    ws3.row_dimensions[r].height = 16
    r += 1

# 合併狀況(1)和(2)的狀況欄
ws3.merge_cells(start_row=4, end_row=5, start_column=1, end_column=1)
ws3.merge_cells(start_row=6, end_row=7, start_column=1, end_column=1)
for row_n in [4, 6]:
    ws3.cell(row=row_n, column=1).alignment = Alignment(horizontal="center", vertical="center")

for i, w in enumerate([6, 36, 4, 10, 12], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4：最小淨間距
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("最小淨間距")

ws4.merge_cells("A1:C1")
ws4["A1"].value = "鋼筋最小淨間距（Minimum Clear Distance）  單位：mm 或 db"
ws4["A1"].font  = Font(size=11, bold=True, name="微軟正黑體", color="FFFFFF")
ws4["A1"].fill  = PatternFill("solid", fgColor="1F4E79")
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 22

clear_headers = ["構件", "方向／條件", "最小淨間距（取大值）"]
for col, h in enumerate(clear_headers, 1):
    c = ws4.cell(row=2, column=col, value=h)
    c.fill = H2_FILL
    c.font = H2_FONT
    c.alignment = H2_ALI
    c.border = THIN_BORDER
ws4.row_dimensions[2].height = 18

clear_data = [
    ("梁（Beam）", "同層水平方向", "1.0 db，且 ≥ 25mm，且 ≥ 骨材最大粒徑 × 4/3"),
    ("",           "垂直（上下層）", "≥ 25mm"),
    ("柱（Column）","柱主筋",        "1.5 db，且 ≥ 40mm，且 ≥ 骨材最大粒徑 × 4/3"),
    ("版（Slab）",  "主筋間距",      "≤ 板厚 × 3，且 ≤ 450mm"),
]
for i, (mem, dir_, dist) in enumerate(clear_data):
    r = 3 + i
    for col, val in enumerate([mem, dir_, dist], 1):
        c = ws4.cell(row=r, column=col, value=val)
        c.font = NORMAL_FONT
        c.alignment = Alignment(horizontal="left" if col > 1 else "center", vertical="center")
        c.border = THIN_BORDER
        c.fill = fills[i % 2]
    ws4.row_dimensions[r].height = 16

ws4.merge_cells("A3:A4")
ws4.merge_cells("A5:A5")
ws4.merge_cells("A6:A6")
ws4["A3"].alignment = Alignment(horizontal="center", vertical="center")

for i, w in enumerate([14, 20, 40], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ── 儲存 ──────────────────────────────────────────────────────────────────────
out = "D:/AI-agent/line-claude-bot/data/regulations/鋼筋混凝土標準規範.xlsx"
wb.save(out)
print(f"已輸出：{out}")
